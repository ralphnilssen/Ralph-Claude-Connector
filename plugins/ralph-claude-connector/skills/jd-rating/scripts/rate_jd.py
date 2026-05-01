"""
rate_jd.py - Generates a scored JD Rating xlsx from a JSON scores payload.

Usage:
    python rate_jd.py <scores_json_path> <output_xlsx_path>
"""
import sys, json, shutil, math
from pathlib import Path

# Checklist rows that contain scored items (excludes section header rows)
CHECKLIST_ROWS = [10,11,13,14,15,16,17,19,20,22,23,24,26,27,28,30,31,33,34,35,36]

# Column E layout constants for row height calculation
COL_E_WIDTH = 55       # Excel width units (matches template)
CHARS_PER_UNIT = 1.2   # approximate chars per width unit for Calibri 11
LINE_HEIGHT_PT = 15    # points per wrapped line
BASE_HEIGHT = 16       # minimum row height in points


def compute_overall_rating(scores: dict) -> str:
    for value in scores.values():
        if value == "Rework":
            return "Not Approved - Reworked"
    return "Approved"


def auto_row_height(text: str) -> float:
    """Estimate row height needed to display wrapped text in column E."""
    if not text:
        return BASE_HEIGHT
    chars_per_line = max(int(COL_E_WIDTH * CHARS_PER_UNIT), 1)
    lines = math.ceil(len(text) / chars_per_line)
    lines = max(lines, 1)
    return max(BASE_HEIGHT, lines * LINE_HEIGHT_PT + 4)


def main():
    if len(sys.argv) < 3:
        print("Usage: rate_jd.py <scores_json> <output_xlsx>")
        sys.exit(1)

    scores_path, output_path = sys.argv[1], sys.argv[2]

    with open(scores_path) as f:
        data = json.load(f)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"], check=True)
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

    template = Path(__file__).parent.parent / "assets" / "JD-Rating-Template.xlsx"
    shutil.copy2(str(template), output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    scores = data.get("scores", {})
    notes = data.get("notes", {})

    # Write status values to D column cells
    for cell, value in scores.items():
        ws[cell] = value

    # Write notes to E column cells with wrap alignment, then set row height
    for row in CHECKLIST_ROWS:
        e_cell_ref = f"E{row}"
        note_text = notes.get(e_cell_ref, "")
        cell = ws[e_cell_ref]
        if note_text:
            cell.value = note_text
        # Always enforce wrap + top-align on E cells
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        # Set row height based on note length
        ws.row_dimensions[row].height = auto_row_height(note_text)

    # Compute and write overall rating to E37
    ws["E37"] = compute_overall_rating(scores)

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
