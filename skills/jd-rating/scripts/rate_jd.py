"""
rate_jd.py - Generates a scored JD Rating xlsx from a JSON scores payload.

Usage:
    python rate_jd.py <scores_json_path> <output_xlsx_path>
"""
import sys, json, shutil
from pathlib import Path


def compute_overall_rating(scores: dict) -> str:
    """Return 'Not Approved - Reworked' if any scored item is Rework, else 'Approved'."""
    for value in scores.values():
        if value == "Rework":
            return "Not Approved - Reworked"
    return "Approved"


def main():
    if len(sys.argv) < 3:
        print("Usage: rate_jd.py <scores_json> <output_xlsx>")
        sys.exit(1)

    scores_path, output_path = sys.argv[1], sys.argv[2]

    with open(scores_path) as f:
        data = json.load(f)

    try:
        from openpyxl import load_workbook
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"], check=True)
        from openpyxl import load_workbook

    template = Path(__file__).parent.parent / "assets" / "JD-Rating-Template.xlsx"
    shutil.copy2(str(template), output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    scores = data.get("scores", {})
    notes = data.get("notes", {})

    # Write status values to D column cells
    for cell, value in scores.items():
        ws[cell] = value

    # Write notes to E column cells
    for cell, note in notes.items():
        ws[cell] = note

    # Compute and write overall rating to E37
    ws["E37"] = compute_overall_rating(scores)

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
